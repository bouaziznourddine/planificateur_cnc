# -*- coding: utf-8 -*-
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
from datetime import datetime, timedelta
import logging
import base64
from io import BytesIO
import base64
from io import BytesIO
from datetime import datetime, time
import logging

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    openpyxl = None

_logger = logging.getLogger(__name__)
try:
    from .genetic_algorithm_scheduler import GeneticAlgorithmScheduler, create_gantt_chart_data
    from .genetic_algorithm_scheduler_helper import create_piece_level_gantt_data
    from .gantt_chart_generator import GanttChartGenerator, generate_statistics_report
    AG_AVAILABLE = True
    print("***********************************             AG ok")
except ImportError:
    AG_AVAILABLE = False
    print("AG non disponible. Installez: pip install plotly pandas numpy")

_logger = logging.getLogger(__name__)


class PlanificateurCNC(models.Model):
    _name = 'planificateur.cnc'
    _description = 'Planificateur CNC avec Algorithme Génétique'
    _order = 'date_creation desc'
    _inherit = ['mail.thread', 'mail.activity.mixin']

    # Informations de base
    nom = fields.Char('Nom du Scénario', required=True, tracking=True)
    date_creation = fields.Datetime(
        'Date Création', default=fields.Datetime.now, readonly=True)
    date_debut = fields.Date('Date Début', required=True, tracking=True)
    date_fin = fields.Date('Date Fin', required=True, tracking=True)
    horizon_planification = fields.Integer(
        'Horizon (jours)', compute='_compute_horizon')

    state = fields.Selection([
        ('draft', 'Brouillon'),
        ('validated', 'Validé'),
        ('optimized', 'Optimisé'),
        ('scheduled', 'Planifié'),
        ('done', 'Terminé'),
        ('cancel', 'Annulé'),
    ], string='État', default='draft', tracking=True)

    # Relations
    machine_ids = fields.Many2many(
        'machine.cnc', string='Machines CNC', required=True)
    of_candidat_ids = fields.Many2many(
        'ordre.fabrication', 'planificateur_of_candidat_rel',
        'planificateur_id', 'of_id', string='OF Candidats'
    )
    of_selectionne_ids = fields.Many2many(
        'ordre.fabrication', 'planificateur_of_selectionne_rel',
        'planificateur_id', 'of_id', string='OF Sélectionnés', readonly=True
    )
    bloc_production_ids = fields.One2many(
        'bloc.production', 'planificateur_id', string='Blocs')
    timeline_ids = fields.One2many(
        'planning.timeline', 'planificateur_id', string='Timeline')

    # Paramètres d'optimisation
    objectif_principal = fields.Selection([
        ('minimize_delays', 'Minimiser les Retards'),
        ('minimize_makespan', 'Minimiser le Makespan'),
        ('maximize_production', 'Maximiser la Production'),
        ('balance_load', 'Équilibrer la Charge'),
    ], string='Objectif', default='minimize_makespan', required=True)

    temps_setup = fields.Integer('Temps Setup (min)', default=30)

    # Paramètres Algorithme Génétique
    use_genetic_algorithm = fields.Boolean('Utiliser AG', default=True)
    ga_population_size = fields.Integer('Population', default=100)
    ga_generations = fields.Integer('Générations', default=200)
    ga_crossover_rate = fields.Float(
        'Taux Croisement', default=0.8, digits=(3, 2))
    ga_mutation_rate = fields.Float(
        'Taux Mutation', default=0.2, digits=(3, 2))

    # Résultats d'optimisation
    makespan_final = fields.Float(
        'Makespan (min)', readonly=True, digits=(16, 2))
    makespan_hours = fields.Float(
        'Makespan (h)', compute='_compute_makespan_hours')
    total_delay = fields.Float(
        'Retard Total (min)', readonly=True, digits=(16, 2))
    machine_balance = fields.Float(
        'Équilibrage', readonly=True, digits=(16, 4))
    optimization_time = fields.Float(
        'Temps Optim (s)', readonly=True, digits=(16, 2))

    # Statistiques
    nb_of_total = fields.Integer(
        'Nombre OF Total', compute='_compute_statistics')
    nb_of_optimises = fields.Integer(
        'Nombre OF Optimisés', compute='_compute_statistics')
    nb_blocs = fields.Integer('Nombre de Blocs', compute='_compute_statistics')
    taux_utilisation = fields.Float(
        'Taux Utilisation (%)', compute='_compute_statistics')

    # Visualisations
    gantt_attachment_id = fields.Many2one(
        'ir.attachment', string="Gantt Attachment", readonly=True)
    convergence_attachment_id = fields.Many2one(
        'ir.attachment', string="Convergence Attachment", readonly=True)
    statistics_report = fields.Text('Statistiques', readonly=True)
    gantt_iframe_html = fields.Html(
        'Gantt Iframe',
        compute='_compute_gantt_iframe',
        sanitize=False
    )
    gantt_by_of_attachment_id = fields.Many2one(
        'ir.attachment', string="Gantt By OF Attachment", readonly=True)
    gantt_by_of_iframe_html = fields.Html(
        'Gantt By OF Iframe',
        compute='_compute_gantt_by_of_iframe',
        sanitize=False
    )
    convergence_iframe_html = fields.Html(
        'Convergence Iframe',
        compute='_compute_convergence_iframe',
        sanitize=False
    )

    @api.depends('gantt_by_of_attachment_id')
    def _compute_gantt_by_of_iframe(self):
        for rec in self:
            if rec.gantt_by_of_attachment_id:
                rec.gantt_by_of_iframe_html = f'''
                    <iframe src="/web/content/{rec.gantt_by_of_attachment_id.id}" 
                            style="width:100%; height:600px; border:none;">
                    </iframe>
                '''
            else:
                rec.gantt_by_of_iframe_html = '<p>Aucun diagramme par OF disponible</p>'

    def action_view_gantt_by_of(self):
        """Ouvrir le Gantt par OF dans un nouvel onglet"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self.gantt_by_of_attachment_id.id}',
            'target': 'new',
        }

    @api.depends('convergence_attachment_id')
    def _compute_convergence_iframe(self):
        for rec in self:
            if rec.convergence_attachment_id:
                rec.convergence_iframe_html = f'''
                    <iframe src="/web/content/{rec.convergence_attachment_id.id}" 
                            style="width:100%; height:400px; border:none;">
                    </iframe>
                '''
            else:
                rec.convergence_iframe_html = '<p>Aucun graphique disponible</p>'

    def action_view_convergence(self):
        """Ouvrir la convergence dans un nouvel onglet"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self.convergence_attachment_id.id}',
            'target': 'new',
        }

    @api.depends('gantt_attachment_id')
    def _compute_gantt_iframe(self):
        for rec in self:
            if rec.gantt_attachment_id:
                rec.gantt_iframe_html = f'''
                        <iframe src="/web/content/{rec.gantt_attachment_id.id}" 
                                style="width:100%; height:600px; border:none;">
                        </iframe>
                    '''
            else:
                rec.gantt_iframe_html = '<p>Aucun diagramme disponible</p>'

    # Export
    excel_file = fields.Binary('Fichier Excel', readonly=True)
    excel_filename = fields.Char('Nom Fichier')

    # Notes
    notes = fields.Text('Notes')

    @api.depends('date_debut', 'date_fin')
    def _compute_horizon(self):
        for rec in self:
            if rec.date_debut and rec.date_fin:
                delta = (rec.date_fin - rec.date_debut).days
                rec.horizon_planification = delta
            else:
                rec.horizon_planification = 0

    @api.depends('makespan_final')
    def _compute_makespan_hours(self):
        for rec in self:
            rec.makespan_hours = rec.makespan_final / 60.0 if rec.makespan_final else 0

    @api.depends('of_candidat_ids', 'of_selectionne_ids', 'bloc_production_ids', 'makespan_final')
    def _compute_statistics(self):
        for rec in self:
            rec.nb_of_total = len(rec.of_candidat_ids)
            rec.nb_of_optimises = len(rec.of_selectionne_ids)
            rec.nb_blocs = len(rec.bloc_production_ids)

            if rec.makespan_final and len(rec.machine_ids) > 0:
                total_capacity = rec.makespan_final * len(rec.machine_ids)
                total_used = sum(
                    b.duree_totale for b in rec.bloc_production_ids)
                rec.taux_utilisation = (
                    total_used / total_capacity * 100) if total_capacity else 0
            else:
                rec.taux_utilisation = 0

    def action_validate(self):
        """Valider le scénario"""
        self.ensure_one()
        if not self.of_candidat_ids:
            raise UserError("Ajoutez au moins un OF candidat.")
        if not self.machine_ids:
            raise UserError("Ajoutez au moins une machine.")
        self.state = 'validated'

    def action_optimiser(self):
        """Optimiser avec AG ou heuristique"""
        self.ensure_one()

        if not self.of_candidat_ids:
            raise UserError("Aucun OF candidat.")
        if not self.machine_ids:
            raise UserError("Aucune machine.")

        if self.use_genetic_algorithm and AG_AVAILABLE:
            print("***********************************            AG ok2")
            return self._optimize_with_ga()

        else:
            print("***********************************            heuristic ok2")
            return self._optimize_heuristic()

        return {
            'type': 'ir.actions.client',
            'tag': 'reload',
        }

    def action_reset(self):
        """Réinitialise le planificateur pour permettre une nouvelle optimisation"""
        self.ensure_one()

        # Supprimer les résultats précédents
        self.write({
            'state': 'validated',  # ou 'draft' selon ton workflow
            'resultat_gantt': False,
            'resultat_convergence': False,
            'resultat_stats': False,
            # Ajouter ici tous les champs de résultats à réinitialiser
            # Par exemple:
            # 'best_makespan': 0,
            # 'best_fitness': 0,
            # 'planning_ids': [(5, 0, 0)],  # Supprimer les lignes de planning
        })

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'current',
        }

    def _optimize_with_ga(self):
        """Optimisation par algorithme génétique"""
        start = datetime.now()

        try:
            tool_capacity = self.machine_ids[0].capacite_magasin or 40

            ga = GeneticAlgorithmScheduler(
                ofs=self.of_candidat_ids,
                machines=self.machine_ids,
                setup_time=self.temps_setup,
                tool_capacity=tool_capacity,
                population_size=self.ga_population_size,
                generations=self.ga_generations,
                crossover_rate=self.ga_crossover_rate,
                mutation_rate=self.ga_mutation_rate,
                objective=self._map_objective()
            )

            solution, stats = ga.run()

            self._apply_solution(solution, ga.of_data)

            self.write({
                'makespan_final': stats['makespan'],
                'total_delay': stats['total_delay'],
                'machine_balance': stats['machine_balance'],
                'optimization_time': (datetime.now() - start).total_seconds(),
                'state': 'optimized',
            })

            self._generate_visualizations(solution, ga.of_data, stats)

            # return {
            #     'type': 'ir.actions.client',
            #     'tag': 'display_notification',
            #     'params': {
            #         'title': 'Optimisation Réussie!',
            #         'message': f'Makespan: {self.makespan_final:.0f} min ({self.makespan_hours:.1f}h)',
            #         'type': 'success',
            #         'sticky': False,
            #     }
            # }
            return {
                'type': 'ir.actions.act_window',
                'res_model': self._name,
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'current',
            }
    #         return {
    #     'type': 'ir.actions.client',
    #     'tag': 'reload',
    # }

        except Exception as e:
            _logger.error(f"Erreur AG: {e}", exc_info=True)
            raise UserError(f"Erreur d'optimisation: {str(e)}")

    def _optimize_heuristic(self):
        """Optimisation heuristique simple"""
        # TODO: Implémenter heuristique
        self.state = 'optimized'
        return True

    def _map_objective(self):
        """Mapper objectif Odoo vers AG"""
        mapping = {
            'minimize_delays': 'delay',
            'minimize_makespan': 'makespan',
            'maximize_production': 'makespan',
            'balance_load': 'balance',
        }
        return mapping.get(self.objectif_principal, 'makespan')

    def _apply_solution(self, solution, of_data):
        """Appliquer la solution AG"""

        self.bloc_production_ids.unlink()
        self.timeline_ids.unlink()

        for idx, block_ofs in enumerate(solution.block_structure):
            machine = self.machine_ids[solution.machine_assignments[idx]]
            total_tools = sum(of_data[of_id]['total_tools']
                              for of_id in block_ofs)
            total_time = sum(of_data[of_id]['total_time']
                             for of_id in block_ofs)

            bloc = self.env['bloc.production'].create({
                'planificateur_id': self.id,
                'nom': f'Bloc {idx + 1}',
                'machine_id': machine.id,
                'sequence': idx + 1,
                'capacite_outils_utilisee': total_tools,
                'duree_totale': total_time,
            })

            ofs = self.of_candidat_ids.filtered(lambda o: o.id in block_ofs)
            for of in ofs:
                of.write(
                    {'bloc_id': bloc.id, 'machine_assignee_id': machine.id, 'state': 'scheduled'})

        self.of_selectionne_ids = [(6, 0, [of.id for of_id in solution.sequence
                                           for of in self.of_candidat_ids.filtered(lambda o: o.id == of_id)])]

    def action_view_gantt(self):
        """Ouvrir le Gantt dans un nouvel onglet"""
        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/{self.gantt_attachment_id.id}',
            'target': 'new',
        }

    def _generate_visualizations(self, solution, of_data, stats):
        """Générer Gantt et graphiques"""
        try:
            gantt_data = create_gantt_chart_data(
                solution, of_data, self.machine_ids,
                self.temps_setup, datetime.combine(
                    self.date_debut, datetime.min.time()) if self.date_debut else datetime.now()
            )

            gen = GanttChartGenerator(gantt_data, f"Planning - {self.nom}")
            fig = gen.generate_advanced_gantt()
            html_content = fig.to_html(include_plotlyjs=True)

            # Créer attachment pour Gantt
            gantt_attachment = self._create_html_attachment(
                f'gantt_{self.id}.html',
                html_content
            )
            self.gantt_attachment_id = gantt_attachment.id

            # Gantt par OF (Niveau Pièce)
            gantt_data_pieces = create_piece_level_gantt_data(
                solution, of_data, self.machine_ids,
                self.temps_setup, datetime.combine(
                    self.date_debut, datetime.min.time()) if self.date_debut else datetime.now()
            )

            gen_of = GanttChartGenerator(
                gantt_data_pieces, f"Planning Détail Pièces - {self.nom}")
            fig_of = gen_of.generate_gantt_by_of()
            html_content_of = fig_of.to_html(include_plotlyjs=True)
            gantt_of_attachment = self._create_html_attachment(
                f'gantt_by_of_{self.id}.html',
                html_content_of
            )
            self.gantt_by_of_attachment_id = gantt_of_attachment.id

            # Convergence chart
            if 'best_fitness_history' in stats:
                fig_conv = GanttChartGenerator.create_convergence_chart(
                    stats['best_fitness_history'], stats['avg_fitness_history']
                )
                conv_html = fig_conv.to_html(include_plotlyjs=True)
                conv_attachment = self._create_html_attachment(
                    f'convergence_{self.id}.html',
                    conv_html
                )
                self.convergence_attachment_id = conv_attachment.id

            report = generate_statistics_report(stats, gantt_data)
            self.statistics_report = self._format_stats(report)

        except Exception as e:
            _logger.error(f"Erreur visualisation: {e}")

    def _create_html_attachment(self, filename, html_content):
        """Créer un attachment HTML"""
        attachment_vals = {
            'name': filename,
            'type': 'binary',
            'datas': base64.b64encode(html_content.encode('utf-8')),
            'res_model': self._name,
            'res_id': self.id,
            'mimetype': 'text/html',
            'public': True,
        }
        return self.env['ir.attachment'].create(attachment_vals)

    def _format_stats(self, report):
        """Formater rapport statistique"""
        text = "="*60 + "\n"
        text += "RAPPORT STATISTIQUE\n"
        text += "="*60 + "\n\n"
        text += f"Makespan: {report['makespan']:.2f} min ({report['makespan']/60:.2f}h)\n"
        text += f"Retard Total: {report['total_delay']:.2f} min\n"
        text += f"Tâches: {report['total_tasks']}\n"
        text += f"Setups: {report['total_setups']}\n\n"

        text += "MACHINES:\n"
        text += "-"*60 + "\n"
        for machine, data in report['machines'].items():
            text += f"{machine}: {data['total_time']:.2f}min, {data['num_tasks']} tâches\n"

        return text

    def action_export_excel(self):
        """Exporter les résultats en Excel avec OpenPyXL - Correction Dates et Mapping"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(
                "La librairie 'openpyxl' n'est pas installée. (pip install openpyxl)")

        # Création du classeur
        wb = openpyxl.Workbook()

        # --- CONFIGURATION DES STYLES ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(
            start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal='center', vertical='center')
        border_style = Side(style='thin')
        thin_border = Border(left=border_style, right=border_style,
                             top=border_style, bottom=border_style)

        # --- FEUILLE 1 : RÉSUMÉ ---
        ws_summary = wb.active
        ws_summary.title = "Résumé"

        summary_data = [
            ["Scénario", self.nom],
            ["Date Début", self.date_debut],
            ["Date Fin", self.date_fin],
            ["Objectif", self.objectif_principal],
            [],
            ["RÉSULTATS", ""],
            ["Makespan (min)", self.makespan_final],
            ["Retard Total (min)", self.total_delay],
            ["Taux Utilisation (%)", f"{self.taux_utilisation:.2f}%"],
            ["Nombre OF", self.nb_of_optimises],
        ]

        for row_idx, row_data in enumerate(summary_data, 1):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_idx, column=col_idx, value=str(
                    value) if value is not None else "")
                if row_idx == 6 or col_idx == 1:
                    cell.font = Font(bold=True)

        ws_summary.column_dimensions['A'].width = 25
        ws_summary.column_dimensions['B'].width = 40

        # --- FEUILLE 2 : DÉTAIL PLANNING ---
        ws_plan = wb.create_sheet("Planning Détaillé")

        headers = [
            'Machine', 'Bloc', 'Séquence', 'Numéro OF',
            'Type Pièce', 'Quantité', 'Date Début', 'Date Fin',
            'Outils Requis', 'Priorité', 'Durée (min)', 'État'
        ]

        # Écriture des en-têtes
        for col_idx, header in enumerate(headers, 1):
            cell = ws_plan.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        # --- MAPPING AVANCÉ DES DATES ---
        # On construit un dictionnaire pour retrouver les dates rapidement
        # Clé : (ID_BLOC, ID_OF) pour unicité
        timeline_map = {}

        # On parcourt la timeline (source de vérité pour les dates)
        for t in self.timeline_ids:
            if t.of_id:  # Si c'est une tâche de production (pas un setup)
                # On stocke les dates. Clé = ID de l'OF
                # Si vous avez le champ bloc_id dans timeline, utilisez (t.bloc_id.id, t.of_id.id)
                # Sinon on utilise juste of_id (attention si un OF est scindé)
                key = t.of_id.id
                if t.bloc_id:
                    key = (t.bloc_id.id, t.of_id.id)

                timeline_map[key] = {
                    'start': t.date_debut,
                    'end': t.date_fin
                }

        # Remplissage des données
        row = 2
        # Tri par Machine puis Séquence pour un affichage logique
        sorted_blocs = self.bloc_production_ids.sorted(
            key=lambda b: (b.machine_id.nom, b.sequence))

        for bloc in sorted_blocs:
            for of in bloc.of_ids:
                # Récupération des dates
                # 1. Essai avec la clé composite (Bloc, OF)
                dates = timeline_map.get((bloc.id, of.id))

                # 2. Si échec, essai avec juste l'ID OF (au cas où le lien bloc manque dans timeline)
                if not dates:
                    dates = timeline_map.get(of.id)

                # Formatage des dates
                start_str = ""
                end_str = ""
                if dates:
                    if dates.get('start'):
                        start_str = dates['start'].strftime('%d/%m/%Y %H:%M')
                    if dates.get('end'):
                        end_str = dates['end'].strftime('%d/%m/%Y %H:%M')

                # Écriture ligne
                ws_plan.cell(row=row, column=1, value=bloc.machine_id.nom)
                ws_plan.cell(row=row, column=2, value=bloc.nom)
                ws_plan.cell(row=row, column=3, value=bloc.sequence)
                ws_plan.cell(row=row, column=4, value=of.numero_of)
                ws_plan.cell(
                    row=row, column=5, value=of.type_piece_id.nom if of.type_piece_id else "")
                ws_plan.cell(row=row, column=6, value=of.quantite)

                # Dates
                ws_plan.cell(row=row, column=7,
                             value=start_str).alignment = center_align
                ws_plan.cell(row=row, column=8,
                             value=end_str).alignment = center_align

                ws_plan.cell(row=row, column=9, value=of.nombre_outils_requis)
                ws_plan.cell(row=row, column=10, value=of.priorite)
                ws_plan.cell(row=row, column=11, value=of.temps_total_estime)

                # État traduit
                state_val = dict(of._fields['state'].selection).get(
                    of.state) if of.state else ""
                ws_plan.cell(row=row, column=12, value=state_val)

                # Bordures
                for col in range(1, 13):
                    ws_plan.cell(row=row, column=col).border = thin_border

                row += 1

        # Ajustement automatique largeurs
        for column_cells in ws_plan.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            ws_plan.column_dimensions[get_column_letter(
                column_cells[0].column)].width = min(length + 3, 50)

        # --- SAUVEGARDE ---
        fp = BytesIO()
        wb.save(fp)
        fp.seek(0)
        data = fp.read()
        fp.close()

        filename = f"Planning_{self.nom.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        self.write({
            'excel_file': base64.b64encode(data),
            'excel_filename': filename
        })

        return {
            'type': 'ir.actions.act_url',
            'url': f'/web/content/?model={self._name}&id={self.id}&field=excel_file&download=true&filename={filename}',
            'target': 'self',
        }

    def action_cancel(self):
        """Annuler"""
        self.state = 'cancel'

    def action_reset_draft(self):
        """Remettre en brouillon"""
        self.write({
            'state': 'draft',
            'makespan_final': 0,
            'total_delay': 0,
            'machine_balance': 0,
            'optimization_time': 0,
            'gantt_chart_html': False,
            'convergence_chart_html': False,
            'statistics_report': False,
        })
        self.bloc_production_ids.unlink()
        self.timeline_ids.unlink()

    @api.constrains('date_debut', 'date_fin')
    def _check_dates(self):
        for rec in self:
            if rec.date_debut and rec.date_fin and rec.date_fin < rec.date_debut:
                raise ValidationError(
                    "La date de fin doit être après la date de début")

    @api.constrains('ga_population_size', 'ga_generations')
    def _check_ga_params(self):
        for rec in self:
            if rec.use_genetic_algorithm:
                if not (10 <= rec.ga_population_size <= 500):
                    raise ValidationError("Population: 10-500")
                if not (10 <= rec.ga_generations <= 1000):
                    raise ValidationError("Générations: 10-1000")

    @api.constrains('ga_crossover_rate', 'ga_mutation_rate')
    def _check_ga_rates(self):
        for rec in self:
            if rec.use_genetic_algorithm:
                if not (0 <= rec.ga_crossover_rate <= 1):
                    raise ValidationError("Croisement: 0-1")
                if not (0 <= rec.ga_mutation_rate <= 1):
                    raise ValidationError("Mutation: 0-1")
